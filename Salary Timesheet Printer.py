#This version will write data from monday.com into an excel timesheet
#template will need to be in C:\Timesheets, and generated folder will be C:\Timesheets\Generated Timesheets
#will specify the person and use the project info
#now gets all info
#anyting that is to characters (ie. dg, dp, f1, f2) is a data frame

#monday imports
import itertools
from typing_extensions import Self, clear_overloads
import requests
import json
from monday import MondayClient
#Excel imports
import asposecells
from xlwt import Workbook
import openpyxl, os
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from PIL import Image
from fileinput import filename
from openpyxl.styles import Alignment
#data processing imports
import pandas as pd
import numpy as np
import datetime
from datetime import datetime
import shutil
from audioop import mul
from operator import index
from sqlite3 import Timestamp
from datatest import validate
#gui imports
import tkinter as tk
from tkinter import END, HORIZONTAL, Listbox, ttk
from tkinter.messagebox import showinfo
from turtle import bgcolor
from click import progressbar
import webbrowser
from tkinter.font import BOLD
from tkinter import *
import sys
#PDF imports
from win32com import client

#create extract function
def json_extract(obj, key):
     """Recursively fetch values from nested JSON."""
     arr = []

     def extract(obj, arr, key):
         """Recursively search for values of key in JSON tree."""
         if isinstance(obj, dict):
              for k, v in obj.items():
                  if isinstance(v, (dict, list)):
                        extract(v, arr, key)
                  elif k == key:
                    arr.append(v)
         elif isinstance(obj, list):
              for item in obj:
                  extract(item, arr, key)
         return arr

     values = extract(obj, arr, key)
     return values

#greyed out entry text function
def handle_focus_in(_):
        if (date_entry.get()=="mm.dd.yyyy"):
            date_entry.delete(0, tk.END)
        date_entry.config(fg='black')
def handle_focus_out(_):
    if len(date_entry.get()) == 0:
        date_entry.delete(0, tk.END)
        date_entry.config(fg='grey')
        date_entry.insert(0, 'mm.dd.yyyy')
def handle_enter(txt):
    handle_focus_out('dummy')

#create requierement function
def strftime_format(format):
    def func(value):
        try:
            datetime.strptime(value,format)
        except ValueError:
            tk.messagebox.showinfo(title = None, message= 'Please enter a valid date in the form yyyy.mm.dd.')
            progress_bar.forget()
            return False
        return True
    func.__doc__=f'should use date format {format}'
    return func

#create link function
def callback(url):
    webbrowser.open_new_tab(url)

#get groups
apiKey = "eyJhbGciOiJIUzI1NiJ9.eyJ0aWQiOjMyNjI0MDE4MiwiYWFpIjoxMSwidWlkIjo0ODQzNjQwOCwiaWFkIjoiMjAyNC0wMi0yN1QxOToyMzo0Ny4wMDBaIiwicGVyIjoibWU6d3JpdGUiLCJhY3RpZCI6MTEzMTkxNTYsInJnbiI6InVzZTEifQ.T9eJPRAhJiEGEL6PdTLA8-9tP-fgoacLZntjSyIsyMk"
apiUrl = "https://api.monday.com/v2"
headers = {"Authorization" : apiKey}

# old key : eyJhbGciOiJIUzI1NiJ9.eyJ0aWQiOjE3MzgwNDIzNSwidWlkIjoyODI5NjY2NSwiaWFkIjoiMjAyMi0wOC0wNFQxMTo0NjozNS4wMDBaIiwicGVyIjoibWU6d3JpdGUiLCJhY3RpZCI6MTEzMTkxNTYsInJnbiI6InVzZTEifQ.bwXbe720qElxhZbILqR7PRrENF0X73hVRR84cCAcGnk

query  = ''' query{ boards (ids:2328990214){ 
    groups{
        title
        id
    } } }'''

data = {'query' : query}
r = requests.post(url=apiUrl, json=data, headers=headers)
output3 = r.json()

#json of group data
group_ids=json_extract(output3, 'id')
group_name=json_extract(output3, 'title')

#df of group ids
gi=pd.DataFrame(group_ids)
gi.columns=['Group Ids']

#df of combined name and id
gn=pd.DataFrame(group_name)

#splits groupnames and employee id
gn2=gn.apply(lambda x: x.str.split('-').explode()).reset_index(drop=True)

#gets just group name
gz=gn2[::2].reset_index(drop=True)
gz.columns=['Group Names']

#gets just employee id
gq=gn2[1::2].reset_index(drop=True)
gq.columns=['Employee Number']

#makes  one dataframe
gh=[gz,gq]
gy=pd.concat(gh, axis=1)

#combined df of group name, employee number and group id
gi2=[gi,gy]
info3 = pd.concat(gi2, axis=1)
num_of_groups=len(info3)

#remove unwanted groups
info4=info3[info3['Group Names'].str.contains(r'[*]')==False]
info5=info4.reset_index(drop=True)

#create gui
root =tk.Tk()
root.title("Monday.com Timesheet Generator")
root.iconbitmap(sys.executable)
root.geometry('750x700')
moss_green='#046a38'
text_color= '#ffffff'
f_preset=('Times',15, 'bold')
root['bg'] = moss_green
message = tk.Label(root, text='Prior to generating a timesheet, make sure you:\n\t-Create the folder C:\Timesheets \n\t-Add the Timesheet Template C:\Timesheets\Timesheet_Template.xlsx \n\t-Also add the CMIC template, C:\Timesheets\CMIC Template.csv \n\t\t-Template can be found at the link below \n\t-Add the folder C:\Timsheets\Generated Timesheets \n\t-All Excel Timesheets are closed\nTo generate a Timesheet select a person(people) and enter a date, then select\n Generate Timesheet. To generate a CMIC, enter a date and select Generate CMIC.',font=f_preset, bg='#046a38',fg=text_color, justify= tk.LEFT)
message.pack()
period_ending=tk.StringVar()
progress=tk.IntVar()

#create widgets
widget=tk.Frame(root)
widget.pack(padx=0, pady=0, expand=True)

widget2=tk.Frame(root)
widget2.config(bg=moss_green)
widget2.pack(padx=0, pady=0, expand=True)

#print message
message4=tk.Label(widget, text='Please select a person to print.',fg=text_color, bg=moss_green, font=f_preset)
message4.pack(fill='x')

#create scrollbar
yscrollbar = tk.Scrollbar(widget)
yscrollbar.pack(side = tk.RIGHT, fill = 'y')

#person multiselect
mult_sel=Listbox(widget, selectmode='multiple', font=(1), justify=tk.CENTER, yscrollcommand = yscrollbar.set)
mult_sel.pack(expand=False, fill='x')

for each_item in range(num_of_groups-1):
    mult_sel.insert(END, info5['Group Names'][each_item])

#enter date message
message2=tk.Label(widget2, text='Please enter the period ending date. (mm.dd.yyyy)', bg=moss_green, fg=text_color, font=f_preset)
message2.pack(fill='x')

#period ending entry box
date_entry=tk.Entry(widget2, textvariable=period_ending, fg='black')
date_entry.insert(0, 'mm.dd.yyyy')

#does data default text
date_entry.bind('<FocusIn>', handle_focus_in)
date_entry.bind("<FocusOut>", handle_focus_out)
date_entry.bind('<Return>', handle_enter)

date_entry.pack(fill='x', expand=True)

#progress bar perentage
style = ttk.Style(root)
# add label in the layout
style.layout('text.Horizontal.TProgressbar', 
             [('Horizontal.Progressbar.trough',
               {'children': [('Horizontal.Progressbar.pbar',
                              {'side': 'left', 'sticky': 'ns'})],
                'sticky': 'nswe'}), 
              ('Horizontal.Progressbar.label', {'sticky': 'nswe'})])
# set initial text
style.configure('text.Horizontal.TProgressbar', text='0 %', anchor='center')

#make progress bar
progress_bar=ttk.Progressbar(widget, style='text.Horizontal.TProgressbar',orient=HORIZONTAL, mode='indeterminate', maximum=100, variable=progress)
progress.set(0)
progress_bar.start()

#comand function
def print_data():
    try:
        #start of salary printer
        #pack progress bar
        progress_bar.pack(fill='x')
        progress.set(0)
        root.update()

        #process groups
        dr=pd.DataFrame(mult_sel.curselection())
        #process date
        per_end=period_ending.get()
        #states if no one is selected
        q=0
        if dr.empty == True:
            tk.messagebox.showinfo(title = None, message= 'Please select a person to print.')
            q=1
            progress_bar.forget()

        #create requierement function
        def strftime_format(format):
            def func(value):
                try:
                    datetime.strptime(value,format)
                except ValueError:
                    tk.messagebox.showinfo(title = None, message= 'Please enter a valid date in the form yyyy.mm.dd.')
                    progress_bar.forget()
                    return False
                return True
            func.__doc__=f'should use date format {format}'
            return func
        
        #states if non-existant date is entered
        validate(per_end, strftime_format('%m.%d.%Y'))

        #back to processing groups
        dr.columns=['Index']
        dl=info5.loc[dr['Index']]
        dl2=dl.reset_index(drop=True)
        kd=dl2['Group Names']
        kl=dl2['Group Ids']
        kl.reset_index(drop=True)
        num_selected=len(kl)
        kn=dl2['Employee Number']
        kn.reset_index(drop=True)

        #create file name
        if num_selected==1:
            i_name=dl['Group Names'].to_string(index=False)
            file_name='TimeSheet '+i_name+' '+ per_end+'.xlsx'
            B=1
        elif num_selected > 1:
            i_name=""
            file_name='TimeSheet '+per_end+'.xlsx'
            B=2
        
        #copy and rename the files. If statements doesn't copy if no name or date is entered
        if q ==0:
            source=r"C:\Timesheets\Timesheet_Template.xlsx"
            destination=r"C:\Timesheets\Generated Timesheets\{}".format(file_name)
            shutil.copyfile(source,destination)
        
        #gets date period
        start_date=pd.to_datetime(per_end,format="%m.%d.%Y")-pd.DateOffset(days=12)
        end_date = pd.to_datetime(per_end, format="%m.%d.%Y")-pd.DateOffset(days=1)

        #write period ending, employee name and emplpoyee id to the spreadsheet
        pe2=pd.DataFrame([period_ending.get()], columns=['PE'])
        pe=pe2.copy(deep=True)
        #reset counter
        c=1
        #start group loop
        while c< (num_selected+1):
            #percentage loading bar
            def increment():
                style.configure('text.Horizontal.TProgressbar', text='{:g} %'.format(c/num_selected*100))  # update label
        
        #get previous interval
            w=c-1
            
            #get sheet name
            if B==1:
                i_n=i_name
            elif B==2:
                i_n=kd[w:c].to_string(index=False)

            #copy sheet
            sheet='Timesheet '+i_n
            wb=load_workbook(destination)
            target=wb['Salary Timesheet']
            wb.copy_worksheet(target).title=sheet
            wb.save(destination)

            #get sheets and workbooks to edit
            wb2=openpyxl.load_workbook(destination)
            wd=wb2.active
            ws=wb2.worksheets[c]
            #format workbook
            wd.sheet_properties.pageSetUpPr.fitToPage = True
            wd.page_setup.orientation= wd.ORIENTATION_LANDSCAPE
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.orientation= ws.ORIENTATION_LANDSCAPE
            #add image
            img=openpyxl.drawing.image.Image("C:\Timesheets\Moss logo.jpg")
            img.height=60
            img.width=240
            img.anchor='AA1'
            ws.add_image(img)
            wb2.save(destination)

            with pd.ExcelWriter(destination, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:  
                pe.to_excel(writer, sheet_name=sheet, startcol=25, startrow=4, index=False, header=False)
                kd[c-1:c].to_excel(writer, sheet_name=sheet, startcol=1, startrow=3, index=False, header=False)
                kn[c-1:c].to_excel(writer, sheet_name=sheet, startcol=1, startrow=4, index=False, header=False)

        #get pulse ids for a single group
            group=kl[w:c]
            gid=group.to_string(index=False)
            variables={'input':gid}

            query  = ''' query($input: [String!]){ boards (ids:2328990214 ){ name
            groups(ids: $input){ 
            items 
            {
            id
            } } } }'''

            data = {'query' : query, 'variables': variables}
            r = requests.post(url=apiUrl, json=data, headers=headers)
            output = r.json()

            #get put pulse ids into dataframe
            need = json_extract(output, 'id')
            pi = pd.DataFrame(need)
            length=(len(pi))
            pi.columns=[' ']

            #create placeholder for empty dataframe
            placeholder= {'Date': ["2020-01-01", "2020-01-02", "2020-01-02", "2020-01-03", "2022-03-15", "2020-01-05", per_end, "2020-01-07", "2020-01-08", "2020-01-09"], 'Time': ["0", "0", "0", "0", "0", "0", "0", "0", "0", "0"]}
            dz=pd.DataFrame(placeholder)

            #stores cost codes
            cost_codes={ 'TLBIM': ["0", "1", "-", "3", "0", "2", "1", "-", "0", "0"], 'TLAD':[ "0", "1", "-", "3", "0", "0", "0", "-", "0", "0"], 'Blank':[ " ", " ", "-", " ", " ", " ", " ", "-", " ", " "]}
            dcc=pd.DataFrame(cost_codes)
            dccc=dcc.transpose(copy=False)

        #reset all counter variables
            n=m=1
            added_rows=0
            #start print loop
            while n<=length:
                skip=0

            #feed in pulse ids to get individual items
                l=n-1
                pulse=pi[l:n]
                pulseid=pulse.to_string(index=False)
                pid=int(float(pulseid))
                variables={'input':pid}

            #gets time and dates
                query  = ''' query($input: [Int!]){ boards (ids:2328990214 ){ name 
                items(ids: $input)  
                {name 
                subitems{
                    name column_values(ids: ["date0","numbers"]){
                    title, text
                    } } 
                    } } }'''

                data = {'query' : query, 'variables': variables}
                r = requests.post(url=apiUrl, json=data, headers=headers)
                output = (r.json())

                #get just subitem values
                need2 = json_extract(output, 'text')
                df = pd.DataFrame(need2)

                #get item names for error codes
                need3=json_extract(output, 'name')
                s_n=pd.DataFrame(need3)

                #Get time
                dj=df[1::2].astype(int)
                dj.reset_index(drop=True, inplace=True)
                if dj.empty == False:
                    dj.columns=['Time']
                elif dj.empty == True:
                    dj=dz['Time']
                    skip=1

                #Get dates
                dg=df[0::2] 
                dg.reset_index(drop=True, inplace=True)
                if dg.empty == False:
                    dg.columns=['Date']
                elif dg.empty == True:
                    dg=dz['Date']
                    skip=1

                #make one dataframe, has all dates and times
                dt=[dg,dj]
                info = pd.concat(dt, axis=1)

                #get specific date range
                date_range=(pd.to_datetime(info['Date']) >= start_date) & (pd.to_datetime(info['Date']) <= end_date)
                fg=info.loc[date_range]
                fg['Time'].astype(int)

                #add dates that are the same together
                a = {'Date': 'first', 'Time': 'sum'}
                f1=fg.groupby(['Date'], as_index=False)['Time'].sum()
            
                #Sort dates by order
                f1["Date"] = pd.to_datetime(f1["Date"])
                f2=f1.sort_values(by= "Date")

                #tell uf if a project has hours
                if f1.empty==True:
                    skip=1
        
                #add in dates that are missing with a zero
                s = pd.date_range(start=start_date, end=end_date)
                f3=f1.set_index('Date').reindex(s).fillna(' ').rename_axis('Date').reset_index()

                #Format dataframe to print propely into excel
                f4=f3.transpose(copy=False)
                f5=f4[1::5]
                f6=f5.iloc[:, 0:5]
                f7=f5.iloc[:, 7:12]
                
                #gets additional information
                query = '''query($input: [Int!]){ boards (ids:2328990214 ){ name 
                items(ids: $input){ 
                    group{ id }
                column_values(ids:["connect_boards", "mirror", "mirror8"]){
                        title, text
                   } } } }'''

                data = {'query' : query, 'variables':variables}
                r = requests.post(url=apiUrl, json=data, headers=headers)
                output2 = (r.json())

                #format json data 
                need2 = json_extract(output2, 'text')
                
                #put into seperate dataframes
                dn = pd.DataFrame(need2)
               
                #get name
                dN=dn[0:1]

                #get B-Number
                db=dn[1:2]

                #Get cost cat
                dc=dn[2:3]
                
                #determines which cost code to print
                d2=dc.to_dict(orient='split')

                if d2['data'] == [['TLBIM']]:
                    d3=(dccc[0:1])
                elif d2['data'] == [['TLAD']]:
                    d3=(dccc[1:2])
                else:
                    d3=(dccc[2:3])

                #format b-number properly
                fn=db.apply(lambda x: x.str.split('').explode()).reset_index(drop=True)
                f9 =fn[1:8].transpose(copy=False)

                #reset overhead variables
                x=y=z=0
                a=0
                #Sort overhead
                fb2=dN.to_dict(orient= 'split')

                if fb2['data']  == [['Sick']]:
                    a=1
                    y=1
                    x=1
                elif fb2['data'] == [['Vacation']]:
                    a=1
                    y=2
                    x=1
                elif fb2['data']  == [['Holiday']]:
                    a=1
                    y=3
                    x=1
                elif fb2['data']  == [['Div 225 Department Overhead']]:
                    a=2
                    z=1
                    x=1
                else:
                    a=3

                #remove blank spaces left by overhead
                if x == 1:
                    m-=1
                
                #dynamically adds rows with proper formatting
                if m>10 and a==3 and skip ==0:
                    #opens workbook
                    workbook=openpyxl.load_workbook(destination)
                    ws=workbook.worksheets[c]
                    #gets merged cells range to move
                    lower_bound="A"+str(24+added_rows)
                    upper_bound="AM"+str(34+added_rows)
                    cell_range = lower_bound +":"+ upper_bound
                    target_range = CellRange(range_string=cell_range)
                    #moves merged cells 
                    merged_cell_range = ws.merged_cells.ranges
                    for merged_cell in merged_cell_range :
                        if merged_cell.issubset(target_range):
                            merged_cell.shift(0,1)
                    #adds cells to missing spots
                    ws.insert_rows(24+added_rows,1)

                    #copies format for added rows
                    sheet1= workbook['Salary Timesheet']
                    sheet2= ws
                    for r in range(23,24):
                        for col in range(1,41):
                            sheet2.cell(row=24+y+added_rows,column=col).value = sheet1.cell(row=r, column=col).value
                            sheet2.cell(row=24+y+added_rows,column=col)._style = sheet1.cell(row=r, column=col)._style

                    #update formula with new rows(normal hours)(total per 2 weeks)
                    sheet2["AD"+str(24+added_rows)] = "=SUM(W{0}:AC{0})".format(24+added_rows)
                    sheet2["AL"+str(24+added_rows)] = "=SUM(AE{0}:AK{0})".format(24+added_rows)
                    sheet2["AM"+str(24+added_rows)] = "=SUM(W{0}:AL{0})/2".format(24+added_rows)
                    
                    #update overhead hours(total per 2 weeks)
                    j=0
                    while j<3:
                        sheet2["AD"+str(26+j+added_rows)] = "=SUM(W{0}:AB{0})".format(26+j+added_rows)
                        sheet2["AL"+str(26+j+added_rows)] = "=SUM(AE{0}:AK{0})".format(26+j+added_rows)
                        sheet2["AM"+str(26+j+added_rows)] = "=SUM(W{0}:AL{0})/2".format(26+j+added_rows)
                        j+=1
                    
                    #make dataframe with column index
                    column_headers = [ "Y", "Z", "AA", "AB", "AC", "AG", "AH", "AI", "AJ" ,"AK"]
                    ci=pd.DataFrame(column_headers, columns=['Col Index'])

                    #update total hours per day
                    k=1
                    while k<11:
                        l=k-1
                        col_in=ci[l:k]
                        col_index=col_in['Col Index'].to_string(index=False)
                        sheet2[str(col_index)+str(30+added_rows)]="=SUM({0}10:{0}{1})".format(col_index,29+added_rows)
                        k+=1

                    #update weekly total
                    sheet2["AD"+str(30+added_rows)] = "=SUM(W{0}:AC{0})".format(30+added_rows)
                    sheet2["AL"+str(30+added_rows)] = "=SUM(W{0}:AC{0})".format(30+added_rows)
                    sheet2["AM"+str(30+added_rows)] = "=SUM(AD{0},AL{0})".format(30+added_rows)
                    #save workbook and progress counter
                    workbook.save(destination)
                    added_rows+=1
  
                #get rows to print to (s row is the holiday row, o row is overhead, row is normal work)
                row=9+m
                s_row=23+y+added_rows
                o_row=8+z
            #print final df to excel
                if skip ==0:
                    with pd.ExcelWriter(destination, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:  
                        if a==1:
                            f6.to_excel(writer, sheet_name=sheet, startcol=24, startrow=s_row, index=False, header=False)
                            f7.to_excel(writer, sheet_name=sheet, startcol=32, startrow=s_row, index=False, header=False)
                            
                        elif a==2:
                            f6.to_excel(writer, sheet_name=sheet, startcol=24, startrow=o_row, index=False, header=False)
                            f7.to_excel(writer, sheet_name=sheet, startcol=32, startrow=o_row, index=False, header=False)
                        
                        elif a==3:
                            f6.to_excel(writer, sheet_name=sheet, startcol=24, startrow=row, index=False, header=False)
                            f7.to_excel(writer, sheet_name=sheet, startcol=32, startrow=row, index=False, header=False)
                            dc.to_excel(writer, sheet_name=sheet, startcol=21, startrow=row, index=False, header=False)
                            dN.to_excel(writer, sheet_name=sheet, startcol=0, startrow=row, index=False, header=False)
                            f9.to_excel(writer, sheet_name=sheet, startcol=1,  startrow=row, index=False,  header=False)
                            d3.to_excel(writer, sheet_name=sheet, startcol=11, startrow=row, index=False, header=False)
                elif skip==1 and x!=1:
                    m-=1

                n+=1
                m+=1
                #progress bar
                progress.set(c/num_selected*100)
                increment()
                root.update()
            c+=1

        #remove template
        wb=load_workbook(destination)
        target=wb['Salary Timesheet']
        wb.remove(target)
        wb.save(destination)

        #convert excel to pdf
        excel = client.Dispatch("Excel.Application")
        sheets = excel.Workbooks.Open(destination)
        sheets.ExportAsFixedFormat(0,destination)
        sheets.Close(True)

        #tells user the program is finished
        progress_bar.pack_forget()
        tk.messagebox.showinfo(title = None, message= 'The data is done printing. You can now close the window.')
        #end of print cycle

    #raises pop up error messages
    except(PermissionError):
        tk.messagebox.showinfo(title = 'Permission Error', message= 'The sheet you wish to print is open. Please close it and rerun the program.')
        progress_bar.pack_forget()
    except(FileNotFoundError):
        tk.messagebox.showinfo(title = 'File Not Found Error', message= 'The template file is not found. Please make sure it exists and is in the proper directory mentioned above. The template can be found at the link below.')
        progress_bar.pack_forget()
    except ValueError as ve:
        empty_group=kd[w::c].to_string(index=False)
        

        #tells the difference between different Value Errors
        if ve.args[0] == 'Length mismatch: Expected axis has 0 elements, new values have 1 elements':
            tk.messagebox.showinfo(title = 'Empty Group', message= 'Please check {0} to make sure there is atleast one project on this group. A group must have atleast one project on it. Please add a project to the group, close the program and try again.'.format(empty_group))
            progress_bar.pack_forget()
        else:
            s_n.columns=['Project Name']
            empty_item=s_n['Project Name'][1::2].to_string(index=False)
            tk.messagebox.showinfo(title = 'Missing Items', message= 'Please check {0} under the group {1} \nto make sure there are no empty subitems as it is likely that. All subitems must have time and date entered. If the issue persists please refer to the confluence.'.format(empty_item, empty_group))#, empty_item))
            progress_bar.pack_forget()
    
def print_cmic():
    try:
        progress_bar.pack(fill='x')
        progress.set(0)
        root.update()
        
        #start of cmic printer
        per_end=period_ending.get()
        #gets date period
        start_date=pd.to_datetime(per_end,format="%m.%d.%Y")-pd.DateOffset(days=12)
        end_date = pd.to_datetime(per_end,format="%m.%d.%Y")-pd.DateOffset(days=1)

        #states if non-existant date is entered
        validate(per_end, strftime_format('%m.%d.%Y'))

        #put date into proper format for file#
        date_doc=pd.to_datetime(per_end, format="%M.%d.%Y").strftime("%M%d%Y")
        
        #make doc #
        document_number="EGBW"+date_doc+".csv"
        doc_num="EGBW"+date_doc

        #copy cmic file
        source=r"C:\Timesheets\CMIC Template.csv"
        destination=r"C:\Timesheets\Generated Timesheets\{}".format(document_number)
        shutil.copyfile(source,destination)
        
        #get pulse ids for a single group
        c=1
        l=0
        while c < (num_of_groups):
             #percentage loading bar
            def increment():
                style.configure('text.Horizontal.TProgressbar', text='{:g} %'.format(c/num_of_groups*100))  # update label
            
            w=c-1
            group=info5["Group Ids"][w:c]
            gid=group.to_string(index=False)
            variables={'input':gid}
            query  = ''' query($input: [String!]){ boards (ids:2328990214 ){ name
            groups(ids: $input){ 
            items 
            {
            id
            } } } }'''
            #print(info5['Group Names'][w:c])
            data = {'query' : query, 'variables': variables}
            r = requests.post(url=apiUrl, json=data, headers=headers)
            output = r.json()

            #get put pulse ids into dataframe
            need = json_extract(output, 'id')
            pi = pd.DataFrame(need)
            length=(len(pi))
            pi.columns=['Pulse Id']
            
            #create placeholder for empty dataframe
            placeholder= {'Date': ["2020-01-01", "2020-01-02", "2020-01-02", "2020-01-03", "2022-03-15", "2020-01-05", per_end, "2020-01-07", "2020-01-08", "2020-01-09"], 'Normal Hours': [0,0,0,0,0,0,0,0,0,0]}
            dz=pd.DataFrame(placeholder)

            #reset all counter variables
            n=1
            #start print loop
            while n<length+1:
                skip=0

            #feed in pulse ids to get individual items
                l=n-1
                pulse=pi[l:n]
                pulseid=pulse.to_string(index=False, header=False)
                pid=int(float(pulseid))
                variables={'input':pid}

            #gets time and dates
                query  = ''' query($input: [Int!]){ boards (ids:2328990214 ){ name 
                items(ids: $input)  
                {name 
                subitems{
                    name column_values(ids: ["date0","numbers"]){
                    title, text
                    } } } } }'''

                data = {'query' : query, 'variables': variables}
                r = requests.post(url=apiUrl, json=data, headers=headers)
                output = (r.json())
                
                #get just subitem values
                need2 = json_extract(output, 'text')
                df = pd.DataFrame(need2)

                #get item names for error codes
                need3=json_extract(output, 'name')
                s_n=pd.DataFrame(need3)

                #Get time
                dh=df[1::2].astype(int)
                dj=dh.reset_index(drop=True)
                if dj.empty == False:
                    dj.columns=['Normal Hours']
                elif dj.empty == True:
                    dj=pd.DataFrame(dz['Normal Hours'])
                    dj.columns=['Normal Hours']
                    skip=1

                #Get dates
                dd=df[0::2] 
                dg=dd.reset_index(drop=True)
                if dg.empty == False:
                    dg.columns=['Date']
                elif dg.empty == True:
                    dg=dz['Date']
                    skip=1
                
                #make one dataframe, has all dates and times
                dt=[dg,dj]
                info = pd.concat(dt, axis=1)
                
                #get specific date range
                date_range=(pd.to_datetime(info['Date']) >= start_date) & (pd.to_datetime(info['Date']) <= end_date)
                fg=info.loc[date_range]
                fg['Normal Hours'].astype(int)

                #add dates that are the same together
                a = {'Date': 'first', 'Normal Hours': 'sum'}
                f1=fg.groupby(['Date'], as_index=False)['Normal Hours'].sum()#this is where normal hours is removed

                #Sort dates by order
                f1["Date"] = pd.to_datetime(f1["Date"], format="%Y-%m-%d")
                f1.sort_values(by= "Date", inplace=True)
                
                #get employee number by itself
                E_Number=info5["Employee Number"][w:c].to_string(index=False, header=False)

                #add employee number to df
                f1.insert(0, "Employee Number", E_Number)
                
                #gets additional information
                query = '''query($input: [Int!]){ boards (ids:2328990214 ){ name 
                items(ids: $input)
                { group{ id }
                column_values(ids:["connect_boards", "mirror", "mirror8"]){
                        title, text
                    } } } }'''

                data = {'query' : query, 'variables':variables}
                r = requests.post(url=apiUrl, json=data, headers=headers)
                output2 = (r.json())

                #format json data 
                need2 = json_extract(output2, 'text')

                #put into seperate dataframes
                do = pd.DataFrame(need2)
                dn=do[1:2]
                dc=do[2:3]

                #formating
                dn.columns=['B']

                #formating
                dc.columns=['Category']
                dc.reset_index(drop=True, inplace=True)

                #determines which cost code to print
                d2=dc.to_dict(orient='split')
                
                if d2['data'] == [['TLBIM']]:
                    cost_code="01.3021.00"
                else:
                    cost_code="01.3000.00"

                #do ot and Double OT(DOT) hours
                OT_Hours="0"

                #do company code
                company_code="01"

                #do other hours
                if d2['data'] == [['VACA']]:
                    Other_Hours=f1['Normal Hours']
                    Other_Hours.to_string(index=False, header=False)
                    
                elif d2['data'] == [['SICK']]:
                    Other_Hours=f1['Normal Hours']
                    Other_Hours.to_string(index=False, header=False)
                    
                else:
                    Other_Hours='0'

                #change date format
                f1["Date"]=f1['Date'].dt.strftime('%d-%b-%Y')
                Blank=" "

                #add extras to f1
                f1.insert(2, "Job #", dn.to_string(index=False, header=False))
                f1.insert(3, "Cost Code",cost_code)
                f1.insert(4, "Cost Code Category",dc.to_string(index=False, header=False))
                f1.insert(6, "OT Hours", OT_Hours)
                f1.insert(7, "Double OT Hours", OT_Hours)
                f1.insert(8, "Other Hours", Other_Hours)
                f1.insert(9, "Cost Code Cat", dc.to_string(index=False, header=False))
                f1.insert(10, "Company Code", company_code)
                f1.insert(11, "S1", Blank)
                f1.insert(12, "S2", Blank)
                f1.insert(13, "S3", Blank)
                f1.insert(14, "S4", Blank)
                f1.insert(15, "Document Number", doc_num)
                
                #change data type
                f1.astype(str)
                f1['Date'] = f1['Date'].apply('="{}"'.format)
                f1['Company Code'] = f1['Company Code'].apply('="{}"'.format)
                
                #prints df to csv
                f1.to_csv(destination, mode='a', index=False, header=False)
                
                #itterate counter
                n+=1

                #progress bar
                progress.set(c/num_of_groups*100)
                increment()
                root.update()

            #iterate counter
            c+=1

        #tells user the program is finished
        progress_bar.pack_forget()
        tk.messagebox.showinfo(title = None, message= 'The data is done printing. You can now close the window.')
        #end of print cycle

    #raises pop up error messages
    except(PermissionError):
        tk.messagebox.showinfo(title = 'Permission Error', message= 'The sheet you wish to print is open. Please close it and rerun the program.')
        progress_bar.pack_forget()
    except(FileNotFoundError):
        tk.messagebox.showinfo(title = 'File Not Found Error', message= 'The template file is not found. Please make sure it exists and is in the proper directory mentioned above. The template can be found at the link below.')
        progress_bar.pack_forget()
    except ValueError as ve:
        empty_group=info5['Group Names'][w:c].to_string(index=False)

        #tells the difference between different Value Errors
        if ve.args[0] == 'Length mismatch: Expected axis has 0 elements, new values have 1 elements':
            tk.messagebox.showinfo(title = 'Empty Group', message= 'Please check {0} to make sure there is atleast one project on this group. A group must have atleast one project on it. Please add a project to the group, close the program and try again.'.format(empty_group))
            progress_bar.pack_forget()
        else:
            s_n.columns=['Project Name']
            empty_item=s_n['Project Name'][2:].to_string(index=False)
            empty_name=s_n['Project Name'][1:2].to_string(index=False)
            
            tk.messagebox.showinfo(title = 'Missing Items', message= 'Please check {0} under the project {1} for {2} to make sure there are no empty subitems as it is likely that. All subitems must have time and date entered. If the issue persists please refer to the confluence.'.format(empty_item, empty_name,empty_group))
            progress_bar.pack_forget()

#enter button
enter_button =tk.Button(widget2, text='Generate Timesheet', command=lambda:print_data(), font=(1) )
enter_button.pack(fill='x',  expand= True, pady=10, padx=0 )

enter_button2 =tk.Button(widget2, text='Generate CMIC', command=lambda:print_cmic(), font=(1) )
enter_button2.pack(fill='x',  expand= True, pady=10, padx=0 )

#help link
message3=tk.Label(root, justify=tk.LEFT, text="If you need help or are having any issues, refer to the full confluence documentation: ",font=f_preset,bg=moss_green, fg=text_color)
message3.pack()
link=tk.Label(root, text= 'Monday.com Timesheet Generator',font=f_preset,bg=moss_green, fg='blue', cursor='hand2')
link.pack()
click_link='https://id.atlassian.com/login?continue=https%3A%2F%2Fid.atlassian.com%2Fjoin%2Fuser-access%3Fresource%3Dari%253Acloud%253Aconfluence%253A%253Asite%252F148f5b7e-0e00-4367-8fc8-b53923b0ad5b%26continue%3Dhttps%253A%252F%252Fmossvdc.atlassian.net%252Fwiki%252Fspaces%252FMV%252Fpages%252F29229057%252FMonday.com%252BTimesheet%252BGenerator&application=confluence'
link.bind('<Button-1>', lambda e:callback(click_link))

root.mainloop()